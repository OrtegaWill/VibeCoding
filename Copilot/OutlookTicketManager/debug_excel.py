#!/usr/bin/env python3
import openpyxl
import os

# Cambiar al directorio del proyecto
os.chdir('/Users/won/Documents/Cognizant/Vibecoding/Copilot/OutlookTicketManager')

print("🔍 ANÁLISIS DETALLADO DEL ARCHIVO EXCEL")
print("=" * 50)

try:
    wb = openpyxl.load_workbook('Insumo/backlog 06 agosto.xlsx')
    sheet = wb.active
    
    print(f"📊 Archivo: backlog 06 agosto.xlsx")
    print(f"📋 Hoja activa: {sheet.title}")
    print(f"📈 Dimensiones: {sheet.max_row} filas x {sheet.max_column} columnas")
    print()
    
    # Análisis fila por fila
    for row_num in range(1, sheet.max_row + 1):
        print(f"🔍 FILA {row_num}:")
        for col_num in range(1, min(14, sheet.max_column + 1)):
            cell = sheet.cell(row=row_num, column=col_num)
            cell_value = cell.value
            
            # Información detallada de cada celda
            value_info = "None" if cell_value is None else f"'{cell_value}' (tipo: {type(cell_value).__name__})"
            
            if col_num == 1:  # ID de incidencia - crítico
                print(f"  ⭐ Col {col_num:2d}: {value_info}")
            else:
                print(f"     Col {col_num:2d}: {value_info}")
        
        # Después de la fila de encabezados, mostrar validación
        if row_num >= 2:
            id_incidencia = sheet.cell(row=row_num, column=1).value
            id_str = str(id_incidencia).strip() if id_incidencia is not None else ""
            is_valid = not (id_str == "" or id_str == "None")
            print(f"  ✅ Validación ID: {'VÁLIDO' if is_valid else '❌ INVÁLIDO'} - ID procesado: '{id_str}'")
        
        print()
    
    # Resumen de validación
    print("📋 RESUMEN DE VALIDACIÓN:")
    valid_rows = 0
    for row_num in range(2, sheet.max_row + 1):
        id_incidencia = sheet.cell(row=row_num, column=1).value
        if id_incidencia is not None:
            id_str = str(id_incidencia).strip()
            if id_str != "" and id_str != "None":
                valid_rows += 1
                
    print(f"✅ Filas válidas encontradas: {valid_rows}/{sheet.max_row - 1}")
    print(f"🎯 Filas que deberían importarse: {valid_rows}")
    
except Exception as e:
    print(f"❌ Error analizando el archivo: {e}")
    import traceback
    traceback.print_exc()
